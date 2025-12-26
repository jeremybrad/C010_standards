#!/usr/bin/env python3
"""
export_standards_pulse.py - Generate Standards Pulse inventory exports

Outputs:
  - 70_evidence/exports/Standards_Pulse.xlsx (multi-sheet workbook)
  - 70_evidence/exports/Standards_Inventory.csv (flat inventory)
  - 20_receipts/<date>_standards_pulse.md (receipt)

Usage:
  python tools/export_standards_pulse.py

Requires: openpyxl (optional - falls back to CSV-only if not available)
"""

import csv
import os
import re
import subprocess
import sys
from datetime import datetime
from pathlib import Path

# Determine repo root (parent of tools/)
REPO_ROOT = Path(__file__).resolve().parent.parent
EXPORTS_DIR = REPO_ROOT / "70_evidence" / "exports"
RECEIPTS_DIR = REPO_ROOT / "20_receipts"

# Try to import openpyxl
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("WARNING: openpyxl not installed - generating CSV only")
    print("  Install with: pip install openpyxl")


def get_git_info() -> dict:
    """Get current git branch and SHA if available."""
    info = {"branch": "unknown", "sha": "unknown"}
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--abbrev-ref", "HEAD"],
            capture_output=True, text=True, cwd=REPO_ROOT
        )
        if result.returncode == 0:
            info["branch"] = result.stdout.strip()

        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            capture_output=True, text=True, cwd=REPO_ROOT
        )
        if result.returncode == 0:
            info["sha"] = result.stdout.strip()
    except Exception:
        pass
    return info


def get_file_info(path: Path) -> dict:
    """Get file metadata."""
    stat = path.stat()
    return {
        "path": str(path.relative_to(REPO_ROOT)),
        "filename": path.name,
        "size_bytes": stat.st_size,
        "modified_time_iso": datetime.fromtimestamp(stat.st_mtime).isoformat(),
    }


def extract_title_from_md(path: Path) -> str:
    """Extract first H1 heading from markdown file."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line.startswith("# "):
                    return line[2:].strip()
                if line and not line.startswith("#"):
                    break  # Stop if we hit content without finding H1
    except Exception:
        pass
    return ""


def extract_version_from_filename(filename: str) -> str:
    """Extract version from filename like docmeta_v1.2.yaml."""
    match = re.search(r"_v(\d+(?:\.\d+)?)", filename)
    return match.group(1) if match else ""


def extract_canon_block(protocol_path: Path) -> str:
    """Extract the '## Canon (per repo)' section from betty_protocol.md."""
    try:
        with open(protocol_path, "r", encoding="utf-8") as f:
            content = f.read()

        # Find Canon section
        match = re.search(
            r"## Canon \(per repo\)\s*\n(.*?)(?=\n## |\Z)",
            content, re.DOTALL
        )
        if match:
            return match.group(0).strip()
    except Exception:
        pass
    return ""


def scan_directory(directory: Path, extensions: list[str] | None = None) -> list[dict]:
    """Scan directory for files with given extensions."""
    items = []
    if not directory.exists():
        return items

    for path in sorted(directory.iterdir()):
        if path.is_file():
            if extensions and path.suffix not in extensions:
                continue
            if path.name.startswith("."):
                continue
            items.append(get_file_info(path))
    return items


def generate_inventory() -> dict:
    """Generate full standards inventory."""
    timestamp = datetime.now().isoformat()
    git_info = get_git_info()

    inventory = {
        "meta": {
            "timestamp": timestamp,
            "git_branch": git_info["branch"],
            "git_sha": git_info["sha"],
        },
        "protocols": [],
        "schemas": [],
        "taxonomies": [],
        "validators": [],
        "scripts": [],
        "canon_snapshot": "",
    }

    # Protocols
    protocols_dir = REPO_ROOT / "protocols"
    for item in scan_directory(protocols_dir, [".md"]):
        item["title"] = extract_title_from_md(protocols_dir / item["filename"])
        item["notes"] = ""
        inventory["protocols"].append(item)

    # Schemas
    schemas_dir = REPO_ROOT / "schemas"
    for item in scan_directory(schemas_dir, [".yaml", ".yml", ".json"]):
        item["version"] = extract_version_from_filename(item["filename"])
        item["notes"] = f"v{item['version']}" if item["version"] else ""
        inventory["schemas"].append(item)

    # Taxonomies
    taxonomies_dir = REPO_ROOT / "taxonomies"
    for item in scan_directory(taxonomies_dir, [".yaml", ".yml"]):
        item["notes"] = ""
        inventory["taxonomies"].append(item)

    # Validators
    validators_dir = REPO_ROOT / "validators"
    for item in scan_directory(validators_dir, [".py"]):
        if item["filename"].startswith("__"):
            continue
        item["notes"] = ""
        inventory["validators"].append(item)

    # Scripts
    scripts_dir = REPO_ROOT / "scripts"
    for item in scan_directory(scripts_dir, [".sh", ".py", ".ps1"]):
        item["notes"] = ""
        inventory["scripts"].append(item)

    # Canon snapshot
    betty_protocol = protocols_dir / "betty_protocol.md"
    if betty_protocol.exists():
        inventory["canon_snapshot"] = extract_canon_block(betty_protocol)

    # Counts
    inventory["meta"]["counts"] = {
        "protocols": len(inventory["protocols"]),
        "schemas": len(inventory["schemas"]),
        "taxonomies": len(inventory["taxonomies"]),
        "validators": len(inventory["validators"]),
        "scripts": len(inventory["scripts"]),
    }

    return inventory


def write_xlsx(inventory: dict, output_path: Path) -> None:
    """Write inventory to Excel workbook with multiple sheets."""
    if not HAS_OPENPYXL:
        return

    wb = Workbook()

    # Header style
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    def add_header_row(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # Sheet 1: RepoMeta
    ws = wb.active
    ws.title = "RepoMeta"
    add_header_row(ws, ["Property", "Value"])
    ws.append(["Timestamp", inventory["meta"]["timestamp"]])
    ws.append(["Git Branch", inventory["meta"]["git_branch"]])
    ws.append(["Git SHA", inventory["meta"]["git_sha"]])
    ws.append(["Protocols Count", inventory["meta"]["counts"]["protocols"]])
    ws.append(["Schemas Count", inventory["meta"]["counts"]["schemas"]])
    ws.append(["Taxonomies Count", inventory["meta"]["counts"]["taxonomies"]])
    ws.append(["Validators Count", inventory["meta"]["counts"]["validators"]])
    ws.append(["Scripts Count", inventory["meta"]["counts"]["scripts"]])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 40

    # Sheet 2: Protocols
    ws = wb.create_sheet("Protocols")
    add_header_row(ws, ["Path", "Title", "Modified", "Size (bytes)"])
    for item in inventory["protocols"]:
        ws.append([item["path"], item.get("title", ""), item["modified_time_iso"], item["size_bytes"]])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 25

    # Sheet 3: Schemas
    ws = wb.create_sheet("Schemas")
    add_header_row(ws, ["Path", "Version", "Modified", "Size (bytes)"])
    for item in inventory["schemas"]:
        ws.append([item["path"], item.get("version", ""), item["modified_time_iso"], item["size_bytes"]])
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["C"].width = 25

    # Sheet 4: Taxonomies
    ws = wb.create_sheet("Taxonomies")
    add_header_row(ws, ["Path", "Modified", "Size (bytes)"])
    for item in inventory["taxonomies"]:
        ws.append([item["path"], item["modified_time_iso"], item["size_bytes"]])
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25

    # Sheet 5: Validators
    ws = wb.create_sheet("Validators")
    add_header_row(ws, ["Path", "Modified", "Size (bytes)"])
    for item in inventory["validators"]:
        ws.append([item["path"], item["modified_time_iso"], item["size_bytes"]])
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 25

    # Sheet 6: Scripts
    ws = wb.create_sheet("Scripts")
    add_header_row(ws, ["Path", "Modified", "Size (bytes)"])
    for item in inventory["scripts"]:
        ws.append([item["path"], item["modified_time_iso"], item["size_bytes"]])
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 25

    # Sheet 7: CanonSnapshot
    ws = wb.create_sheet("CanonSnapshot")
    add_header_row(ws, ["Canon Block (from protocols/betty_protocol.md)"])
    for line in inventory["canon_snapshot"].split("\n"):
        ws.append([line])
    ws.column_dimensions["A"].width = 100

    wb.save(output_path)
    print(f"  Written: {output_path}")


def write_csv(inventory: dict, output_path: Path) -> None:
    """Write flattened inventory to CSV."""
    rows = []

    for item in inventory["protocols"]:
        rows.append({
            "category": "protocol",
            "path": item["path"],
            "filename": item["filename"],
            "size_bytes": item["size_bytes"],
            "modified_time_iso": item["modified_time_iso"],
            "notes": item.get("title", ""),
        })

    for item in inventory["schemas"]:
        rows.append({
            "category": "schema",
            "path": item["path"],
            "filename": item["filename"],
            "size_bytes": item["size_bytes"],
            "modified_time_iso": item["modified_time_iso"],
            "notes": item.get("notes", ""),
        })

    for item in inventory["taxonomies"]:
        rows.append({
            "category": "taxonomy",
            "path": item["path"],
            "filename": item["filename"],
            "size_bytes": item["size_bytes"],
            "modified_time_iso": item["modified_time_iso"],
            "notes": "",
        })

    for item in inventory["validators"]:
        rows.append({
            "category": "validator",
            "path": item["path"],
            "filename": item["filename"],
            "size_bytes": item["size_bytes"],
            "modified_time_iso": item["modified_time_iso"],
            "notes": "",
        })

    for item in inventory["scripts"]:
        rows.append({
            "category": "script",
            "path": item["path"],
            "filename": item["filename"],
            "size_bytes": item["size_bytes"],
            "modified_time_iso": item["modified_time_iso"],
            "notes": "",
        })

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "category", "path", "filename", "size_bytes", "modified_time_iso", "notes"
        ])
        writer.writeheader()
        writer.writerows(rows)

    print(f"  Written: {output_path}")
    return len(rows)


def write_receipt(inventory: dict, xlsx_written: bool, csv_rows: int, exceptions: list[str]) -> Path:
    """Write receipt to 20_receipts/."""
    date_str = datetime.now().strftime("%Y-%m-%d")
    receipt_path = RECEIPTS_DIR / f"{date_str}_standards_pulse.md"

    content = f"""# Standards Pulse Receipt

**Generated:** {inventory["meta"]["timestamp"]}
**Git Branch:** {inventory["meta"]["git_branch"]}
**Git SHA:** {inventory["meta"]["git_sha"]}

## Output Files

| File | Status |
|------|--------|
| `70_evidence/exports/Standards_Pulse.xlsx` | {"Written" if xlsx_written else "Skipped (openpyxl not installed)"} |
| `70_evidence/exports/Standards_Inventory.csv` | Written ({csv_rows} rows) |

## Row Counts by Category

| Category | Count |
|----------|-------|
| Protocols | {inventory["meta"]["counts"]["protocols"]} |
| Schemas | {inventory["meta"]["counts"]["schemas"]} |
| Taxonomies | {inventory["meta"]["counts"]["taxonomies"]} |
| Validators | {inventory["meta"]["counts"]["validators"]} |
| Scripts | {inventory["meta"]["counts"]["scripts"]} |
| **Total** | **{csv_rows}** |

## Exceptions

{"None" if not exceptions else chr(10).join(f"- {e}" for e in exceptions)}

---
*Generated by `tools/export_standards_pulse.py`*
"""

    with open(receipt_path, "w", encoding="utf-8") as f:
        f.write(content)

    print(f"  Written: {receipt_path}")
    return receipt_path


def main() -> int:
    """Main entry point."""
    print("Generating Standards Pulse inventory...")
    print(f"  Repo root: {REPO_ROOT}")
    print("")

    # Ensure output directories exist
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)

    # Generate inventory
    inventory = generate_inventory()
    exceptions = []

    # Write outputs
    xlsx_path = EXPORTS_DIR / "Standards_Pulse.xlsx"
    csv_path = EXPORTS_DIR / "Standards_Inventory.csv"

    xlsx_written = False
    if HAS_OPENPYXL:
        try:
            write_xlsx(inventory, xlsx_path)
            xlsx_written = True
        except Exception as e:
            exceptions.append(f"XLSX write failed: {e}")
            print(f"  ERROR writing XLSX: {e}")
    else:
        exceptions.append("openpyxl not installed - XLSX skipped")

    csv_rows = write_csv(inventory, csv_path)

    # Write receipt
    write_receipt(inventory, xlsx_written, csv_rows, exceptions)

    print("")
    print("Standards Pulse complete!")
    print(f"  Total items: {csv_rows}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
